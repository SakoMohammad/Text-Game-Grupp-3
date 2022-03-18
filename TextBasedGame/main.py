from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

import re

workbook = load_workbook("Platsnamn textspel.xlsx")
sheet = workbook['Rum']
GameMap = workbook['Karta']
wsSave = workbook['Sparande']

room_description = 1  # for room positions


# sheet.append(["ID",'Room','Room Description', 'Paths'])


def goThroughSheet(thesheet):  # a sheet from the workbook
    # goes through a sheet and makes a list of x positions for every y position, returns a list of lists that can be read as a 2d grid
    main_map = []
    for row in thesheet:
        column_rooms = []  # list of x values
        for column in row:
            column_rooms.append(str(column.value))  # add x value to list
        main_map.append(column_rooms)
    return main_map


def look(room_value):  # list
    # takes the current room type you are in and gives the description of that room type
    print(str(room_value[room_description]))


def save(x, y):  # ints
    # takes the x,y positions and saves them into their own sheet
    wsSave['A1'] = x
    wsSave['A2'] = y
    workbook.save("Platsnamn textspel.xlsx")
    print('Your game progress has now been saved')


def move(x, y, direction, valid_directions):  # int, int,string,list
    # the movement function, takes the current position, checks which direction you want to move, checks if it's possible and returns the new position.

    direction=direction.capitalize()  # makes sure the string is in the right format
    if direction == 'North' and direction in valid_directions:  # if it has the right name and is in the list of availible directions
        y -= 1
    elif direction == 'South' and direction in valid_directions:
        y += 1
    elif direction == 'East' and direction in valid_directions:
        x += 1
    elif direction == 'West' and direction in valid_directions:
        x -= 1
    else:
        print('Enter a valid direction')
    return x, y


# print(goThroughSheet(GameMap))

def main():  # where the game is actually put together
    x = int(wsSave['A1'])  # the starting x value
    y = int(wsSave['A2'])  # the starting y value
    # some starting values to load things into memory
    room_description = 1
    running = True
    room_IDs = goThroughSheet(sheet)  # takes the sheet with room ids and puts them in a list
    main_map = goThroughSheet(GameMap)  # takes the map from the map sheet and puts them in a
    # these 2 are below are going to change in the loop
    player_position = main_map[y][x]  # the player position is based on the mainmap
    current_room_type = room_IDs[int(float(player_position))]  # takes the position

    while running:  # the main game loop where you will be making decisions
        possible_directions = current_room_type[2].split(" ")  # the string with possible directions is split in a list

        a = input()
        if a == 'look':
            look(current_room_type)

        if a == 'save':
            save(x, y)

        if re.search('move ', a):
            text = a
            text = re.sub('move ', '', text)

            x, y = move(x, y, text, possible_directions)
            print(text)
        if a == 'quit':
            running = False

        player_position = main_map[y][x]
        current_room_type = room_IDs[int(float(player_position))]
        # print(player_position)

        # print(str(x) + ' ' + str(y))
        # print(str(current_room_type))


main()
# sheet.append(["1",'hallway','you see a brightly lit hallway infront of you', 'North,South'])

# workbook.save("hello_world.xlsx")
